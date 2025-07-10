from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory
import os
import openpyxl
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'supersecretkey'

HONEY_TYPES = [
    {
        'slug': 'acacia',
        'name': "Акацієвий мед",
        'desc': "Світлий, ніжний, з легким ароматом акації. Дуже корисний для дітей.",
        'img': 'acacia.jpg',
        'price': 375
    },
    {
        'slug': 'buckwheat',
        'name': "Гречаний мед",
        'desc': "Темний, насичений смак, багатий на залізо. Рекомендується при анемії.",
        'img': 'buckwheat.jpg',
        'price': 325
    },
    {
        'slug': 'sunflower',
        'name': "Соняшниковий мед",
        'desc': "Яскравий, золотистий, з приємною кислинкою. Дуже популярний в Україні.",
        'img': 'sunflower.jpg',
        'price': 300
    },
    {
        'slug': 'herbs',
        'name': "Мед різнотрав'я",
        'desc': "Ароматний, зібраний з різних польових квітів. Має широкий спектр корисних властивостей.",
        'img': 'herbs.jpg',
        'price': 400
    },
]

ORDERS_DIR = 'orders'
EXCEL_FILE = os.path.join(ORDERS_DIR, 'orders.xlsx')

def save_orders_to_excel(cart):
    if not cart:
        return
    os.makedirs(ORDERS_DIR, exist_ok=True)
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Мед', 'Ціна за 1 л (грн)', 'Кількість (л)', 'Сума за мед (грн)', "Ім'я", 'Телефон', 'Адреса нової пошти', 'Загальна сума замовлення (грн)'])
    # Підрахунок загальної суми для замовлення
    total = 0
    rows = []
    for item in cart:
        price = 0
        if item.get('honey') == 'Акацієвий мед':
            price = 375
        elif item.get('honey') == 'Гречаний мед':
            price = 325
        elif item.get('honey') == 'Соняшниковий мед':
            price = 300
        elif item.get('honey') == "Мед різнотрав'я":
            price = 400
        amount = float(item.get('amount', 0))
        sum_for_honey = price * amount
        total += sum_for_honey
        rows.append([
            item.get('honey', ''),
            price,
            amount,
            round(sum_for_honey, 2),
            item.get('name', ''),
            item.get('phone', ''),
            item.get('address', ''),
            ''  # Загальна сума буде тільки в останньому рядку
        ])
    # Додаємо всі рядки
    for i, row in enumerate(rows):
        if i == len(rows) - 1:
            row[-1] = round(total, 2)
        ws.append(row)
    # Додаємо порожній рядок для розділення замовлень
    ws.append([''] * len(rows[0]))
    wb.save(EXCEL_FILE)

@app.route('/')
def home():
    return render_template('index.html', honeys=HONEY_TYPES)

@app.route('/honey/<slug>', methods=['GET', 'POST'])
def honey_page(slug):
    honey = next((h for h in HONEY_TYPES if h['slug'] == slug), None)
    if not honey:
        return "Мед не знайдено", 404
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        address = request.form.get('address')
        amount = request.form.get('amount')
        if not name or not phone or not amount or not address:
            flash('Будь ласка, заповніть всі поля!', 'danger')
        else:
            cart = session.get('cart', [])
            cart.append({'honey': honey['name'], 'amount': amount, 'name': name, 'phone': phone, 'address': address})
            session['cart'] = cart
            flash('Товар додано до корзини!', 'success')
            return redirect(url_for('cart'))
    return render_template('honey.html', honey=honey)

@app.route('/cart', methods=['GET'])
def cart():
    cart = session.get('cart', [])
    return render_template('cart.html', cart=cart)

@app.route('/place_order', methods=['POST'])
def place_order():
    cart = session.get('cart', [])
    if not cart:
        flash('Корзина порожня!', 'danger')
        return redirect(url_for('cart'))
    save_orders_to_excel(cart)
    session['cart'] = []
    flash('Замовлення оформлено та збережено у Excel!', 'success')
    return redirect(url_for('cart'))

@app.route('/clear_cart')
def clear_cart():
    session['cart'] = []
    flash('Корзину очищено!', 'success')
    return redirect(url_for('cart'))

@app.route('/media/<path:filename>')
def media(filename):
    return send_from_directory('media', filename)

if __name__ == '__main__':
    app.run(debug=True) 